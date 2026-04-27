"""
Layer-2 deidentification for the ICU-SLEEP source XLSX, preserving cell colors,
fills, fonts, comments, and merged cells. Produces an XLSX with the same data
content as the public CSVs (per-SID date-shifted, DOB blanked, ages floored
and top-coded at 90, free-text Re-admit date redacted), suitable for upload
alongside the CSVs in the public S3 release.

Why XLSX in addition to CSV: the source workbook has color formatting and cell
comments that convey study-team annotations and would be lost in CSV.

Inputs:
  --input PATH    source XLSX (PHI-protected; pulled from Box, not in repo)
  --shifts PATH   per-SID shift table written by deidentify.R
  --output PATH   destination XLSX

Run examples:
  python3 deidentify_xlsx.py \
      --input  /tmp/ICU-SLEEP_Data_P1_Deidentified_11.26.25.xlsx \
      --shifts /tmp/_INTERNAL_DO_NOT_PUBLISH_shifts.csv \
      --output /tmp/ICU-SLEEP_Data_Deidentified_PUBLIC.xlsx
"""
import argparse
import csv
import datetime as dt
import re
import sys

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

EXCEL_EPOCH = dt.datetime(1899, 12, 30)
EXCEL_SERIAL_MIN = 25569   # 1970-01-01
EXCEL_SERIAL_MAX = 80000   # ~2118; "duration" cells are ~0..1 (1899) so excluded


def is_serial_date(v):
    if not isinstance(v, (int, float)):
        return False
    return EXCEL_SERIAL_MIN <= v <= EXCEL_SERIAL_MAX


def serial_to_dt(v):
    return EXCEL_EPOCH + dt.timedelta(days=float(v))


def dt_to_serial(d):
    return (d - EXCEL_EPOCH).total_seconds() / 86400.0


# Column-name patterns that indicate a date column whose contents (datetime
# objects or Excel-serial floats stored as numbers/strings) should be shifted.
DATE_NAME_RX = re.compile(
    r"date|time|dts|admission|discharge|birth|on-study\s+date|"
    r"event_time|drug.*time|withdrawal|iqcode-sf\s+performed",
    re.IGNORECASE,
)
DOB_NAME_RX = re.compile(r"^\s*DOB\s*$", re.IGNORECASE)
AGE_NAME_RX = re.compile(r"^\s*Age\s*\(years\)\s*\[at enrollment\]\s*$", re.IGNORECASE)
READMIT_NAME_RX = re.compile(r"re-?admit\s*date", re.IGNORECASE)
SID_NAME_RX = re.compile(r"^SID(\.\.\.)?", re.IGNORECASE)
DURATION_NAME_RX = re.compile(r"duration", re.IGNORECASE)


def find_header_row(sheet):
    """Heuristic: header is the first non-empty row."""
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(r, c).value
            if v is not None and str(v).strip():
                return r
    return 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--shifts", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    # Load shifts
    shifts = {}
    with open(args.shifts) as f:
        for row in csv.DictReader(f):
            shifts[row["SID"]] = int(row["shift_days"])
    print(f"Loaded {len(shifts)} per-SID shifts.", flush=True)

    # Load workbook (keep_vba=False; preserves formatting, comments, merged cells)
    print("Loading source workbook (this can take ~30s for a 170 MB file)...", flush=True)
    wb = load_workbook(args.input, data_only=False, keep_links=False)

    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        nrow = sh.max_row
        ncol = sh.max_column
        print(f"\n--- {sheet_name}  ({nrow} rows x {ncol} cols) ---", flush=True)

        header_row = find_header_row(sh)
        print(f"  header at row {header_row}")

        headers = [str(sh.cell(header_row, c).value or "") for c in range(1, ncol + 1)]

        # locate columns
        sid_col_idx = None
        for i, h in enumerate(headers, start=1):
            if SID_NAME_RX.match(h):
                sid_col_idx = i
                break
        if sid_col_idx is None:
            print(f"  WARN: no SID column; skipping date shift on this sheet")
        else:
            print(f"  SID col at index {sid_col_idx}: {headers[sid_col_idx-1]!r}")

        date_col_idxs = []
        for i, h in enumerate(headers, start=1):
            if DURATION_NAME_RX.search(h):
                continue  # skip "Duration (hh:mm)..." columns (Excel time-only)
            if DATE_NAME_RX.search(h):
                date_col_idxs.append(i)
        print(f"  {len(date_col_idxs)} date-named columns to shift")

        dob_idxs = [i for i, h in enumerate(headers, start=1) if DOB_NAME_RX.match(h)]
        age_idxs = [i for i, h in enumerate(headers, start=1) if AGE_NAME_RX.match(h)]
        readmit_idxs = [i for i, h in enumerate(headers, start=1) if READMIT_NAME_RX.search(h)]

        n_shifted = n_dob_blanked = n_age_topcoded = n_age_floored = n_readmit_redacted = 0

        # Iterate data rows
        for r in range(header_row + 1, nrow + 1):
            sid = sh.cell(r, sid_col_idx).value if sid_col_idx else None
            sid_str = str(sid) if sid is not None else None
            shift_days = shifts.get(sid_str, 0) if sid_str else 0

            # Date columns
            if shift_days:
                td = dt.timedelta(days=shift_days)
                for c in date_col_idxs:
                    cell = sh.cell(r, c)
                    v = cell.value
                    if v is None:
                        continue
                    if isinstance(v, dt.datetime):
                        cell.value = v + td
                        n_shifted += 1
                    elif isinstance(v, dt.date):
                        cell.value = v + td
                        n_shifted += 1
                    elif isinstance(v, (int, float)) and is_serial_date(v):
                        d = serial_to_dt(v)
                        cell.value = dt_to_serial(d + td)
                        n_shifted += 1
                    elif isinstance(v, str):
                        # Try parsing as Excel serial in string form
                        try:
                            f = float(v)
                            if is_serial_date(f):
                                d = serial_to_dt(f)
                                cell.value = dt_to_serial(d + td)
                                n_shifted += 1
                                continue
                        except (ValueError, TypeError):
                            pass
                        # Free text dates left alone (Re-admit handled below)

            # DOB: blank
            for c in dob_idxs:
                cell = sh.cell(r, c)
                if cell.value is not None:
                    cell.value = None
                    n_dob_blanked += 1

            # Age: floor numeric, top-code 90+ to 90, preserve "90 or above" strings as 90
            for c in age_idxs:
                cell = sh.cell(r, c)
                v = cell.value
                if v is None:
                    continue
                # numeric or numeric-string
                try:
                    fnum = float(v) if not isinstance(v, str) else float(v)
                    if fnum >= 90:
                        cell.value = 90
                        n_age_topcoded += 1
                    else:
                        cell.value = int(fnum)  # floor (drops partial-year)
                        n_age_floored += 1
                except (ValueError, TypeError):
                    # Non-numeric string (e.g. "90 or above; [Redacted]")
                    if isinstance(v, str) and "90" in v:
                        cell.value = 90
                        n_age_topcoded += 1
                    # else leave alone (NA / blank)

            # Re-admit date free text: redact
            for c in readmit_idxs:
                cell = sh.cell(r, c)
                if cell.value is not None:
                    cell.value = "[Date redacted]"
                    n_readmit_redacted += 1

        print(f"  shifted {n_shifted} date cells")
        print(f"  blanked {n_dob_blanked} DOB cells")
        print(f"  age-floored {n_age_floored}, age-topcoded {n_age_topcoded}")
        print(f"  redacted {n_readmit_redacted} re-admit cells")

    print(f"\nWriting output workbook to: {args.output}", flush=True)
    wb.save(args.output)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
